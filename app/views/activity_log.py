# logs/views.py
from django.core.paginator import Paginator
from django.contrib.auth.decorators import login_required, permission_required
from django.shortcuts import render
from ..models import ActivityLog
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from app.utils.activity_log import log_activity 
from django.http import HttpResponse
from django.utils.dateparse import parse_date
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

@login_required(login_url='login')
def activity_log_list(request):
    logs = ActivityLog.objects.select_related('user').all()

    # ===== Filters =====
    username = request.GET.get('username')
    action = request.GET.get('action')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    if username:
        logs = logs.filter(user__username__icontains=username)

    if action:
        logs = logs.filter(action__icontains=action)

    if start_date:
        logs = logs.filter(created_at__date__gte=start_date)

    if end_date:
        logs = logs.filter(created_at__date__lte=end_date)

    # ===== Pagination =====
    paginator = Paginator(logs, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'logs/activity_log_list.html', {
        'page_obj': page_obj
    })


@login_required
def delete_log(request, pk):
    log = get_object_or_404(ActivityLog, pk=pk)
    log.delete()

    # Meta-log: record deletion
    log_activity(
        request,
        action="DELETE",
        model_name="ActivityLog",
        object_id=pk,
        description=f"Deleted audit log entry #{pk}"
    )

    messages.success(request, "Audit log entry deleted successfully.")
    return redirect('activity_log')

@login_required
def export_auditlog_excel(request):

    # Get filter parameters
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    logs = ActivityLog.objects.all()

    if start_date:
        logs = logs.filter(created_at__date__gte=start_date)
    if end_date:
        logs = logs.filter(created_at__date__lte=end_date)

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Audit Logs"

    # ===============================
    # MERGED HEADER AREA
    # ===============================

    TOTAL_COLUMNS = 6  # Username, Action, Model, Description, IP, Date

    # Row 1 — Main Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=TOTAL_COLUMNS)
    ws["A1"] = "Activity Log of Expense Tracking"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    # Row 2 — Generated Date
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=TOTAL_COLUMNS)
    ws["A2"] = f"Generated on: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws["A2"].font = Font(bold=True, size=12)
    ws["A2"].alignment = Alignment(horizontal="left")

    # Row 3 — User
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=TOTAL_COLUMNS)
    ws["A3"] = f"User: {request.user.username}"
    ws["A3"].font = Font(bold=True, size=12)
    ws["A3"].alignment = Alignment(horizontal="left")

    # Row 4 — Filters
    start_label = start_date if start_date else "All"
    end_label = end_date if end_date else "All"

    ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=TOTAL_COLUMNS)
    ws["A4"] = f"Filters → Date: {start_label} to {end_label}"
    ws["A4"].font = Font(bold=True, size=12)
    ws["A4"].alignment = Alignment(horizontal="left")

    # Space before table
    TABLE_START = 6

    # ===============================
    # TABLE HEADERS
    # ===============================

    headers = ["Username", "Action", "Model", "Description", "IP", "Date"]
    for col_num, title in enumerate(headers, start=1):
        cell = ws.cell(row=TABLE_START, column=col_num, value=title)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="C6E0B4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")

    # ===============================
    # TABLE DATA
    # ===============================

    row = TABLE_START + 1

    for log in logs:
        ws.cell(row=row, column=1, value=log.user.username if log.user else "System")
        ws.cell(row=row, column=2, value=log.action)
        ws.cell(row=row, column=3, value=log.model_name)
        ws.cell(row=row, column=4, value=log.description)
        ws.cell(row=row, column=5, value=log.ip_address)
        ws.cell(row=row, column=6, value=log.created_at.strftime("%Y-%m-%d %H:%M:%S"))
        row += 1

    # ===============================
    # RETURN RESPONSE
    # ===============================

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="audit_logs.xlsx"'

    wb.save(response)
    return response

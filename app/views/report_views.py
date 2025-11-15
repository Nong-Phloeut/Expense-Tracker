import openpyxl
from django.http import HttpResponse
from django.core.paginator import Paginator
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.db.models import Sum
from datetime import date
from ..models import Expense, Category
from django.utils import timezone
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter

@login_required(login_url='login')
def reports(request):
    start_date = request.GET.get('startDate')
    end_date = request.GET.get('endDate')
    category_filter = request.GET.get('category')

    expenses = Expense.objects.filter(user=request.user)

   # Default: show current month's expenses if no date range is provided
    today = date.today()
    if not start_date and not end_date:
        expenses = expenses.filter(date__year=today.year, date__month=today.month)
    else:
        if start_date:
            expenses = expenses.filter(date__gte=start_date)
        if end_date:
            expenses = expenses.filter(date__lte=end_date)

    # Apply category filter
    if category_filter:
        expenses = expenses.filter(category__name=category_filter)

   # Apply pagination — 10 per page
    paginator = Paginator(expenses.order_by('-date', '-id'), 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Summary calculations
    total_expenses = expenses.aggregate(Sum('amount'))['amount__sum'] or 0
    highest_expense = expenses.order_by('-amount').first()
    avg_per_day = 0

    if expenses.exists():
        first_date = expenses.order_by('date').first().date
        last_date = expenses.order_by('-date').first().date
        days = (last_date - first_date).days + 1
        avg_per_day = round(total_expenses / days, 2) if days > 0 else total_expenses

    top_category = (
        expenses.values('category__name')
        .annotate(total=Sum('amount'))
        .order_by('-total')
        .first()
    )

    top_category_name = top_category['category__name'] if top_category else "N/A"

    # Data for Chart.js
    category_data = list(
        expenses.values('category__name')
        .annotate(total=Sum('amount'))
        .order_by('-total')
    )
    category_labels = [c['category__name'] for c in category_data]
    category_totals = [float(c['total']) for c in category_data]

    # Monthly trend chart
    monthly_data = (
        expenses.extra(select={'month': "DATE_TRUNC('month', date)"})
        .values('month')
        .annotate(total=Sum('amount'))
        .order_by('month')
    )
    months = [m['month'].strftime('%b %Y') for m in monthly_data]
    monthly_totals = [float(m['total']) for m in monthly_data]

    filters = request.GET.copy()
    # Remove page param so you don't duplicate it
    if 'page' in filters:
        filters.pop('page')

    return render(request, 'reports/reports.html', {
        'expenses': page_obj,
        'total_expenses': total_expenses,
        'highest_expense': highest_expense,
        'avg_per_day': avg_per_day,
        'top_category': top_category_name,
        'category_labels': category_labels,
        'category_totals': category_totals,
        'months': months,
        'monthly_totals': monthly_totals,
        'categories': Category.objects.all(),
        'start_date': start_date or '',
        'end_date': end_date or '',
        'category_filter': category_filter or '',
        'filters': filters.urlencode(),
    })

@login_required(login_url='login')
def export_expenses_excel(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    category = request.GET.get('category')

    expenses = Expense.objects.filter(user=request.user)

    if start_date:
        expenses = expenses.filter(date__gte=start_date)
    if end_date:
        expenses = expenses.filter(date__lte=end_date)
    if category and category != '':
        expenses = expenses.filter(category__name=category)

    # Create workbook and sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Expenses Report"

    # === Styles ===
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    left_alignment = Alignment(horizontal="left", vertical="center")
    border_style = Side(border_style="thin", color="CCCCCC")
    border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)

    # === Report Title ===
    title = "Expenses Report"
    ws.merge_cells("A1:D1")
    title_cell = ws["A1"]
    title_cell.value = title
    title_cell.font = Font(size=14, bold=True)
    title_cell.alignment = center_alignment

    # === Metadata (merged cells for better layout) ===
    ws.merge_cells("A2:D2")
    ws["A2"].value = f"Generated on: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws["A2"].alignment = left_alignment

    ws.merge_cells("A3:D3")
    ws["A3"].value = f"User: {request.user.username}"
    ws["A3"].alignment = left_alignment

    ws.merge_cells("A4:D4")
    ws["A4"].value = f"Filters → Date: {start_date or 'All'} to {end_date or 'All'}"
    ws["A4"].alignment = left_alignment

    # === Blank row before header ===
    ws.append([])

    # === Header Row ===
    headers = ["Date", "Category", "Description", "Amount (USD)"]
    ws.append(headers)
    header_row = ws[6]  # after title & metadata

    for cell in header_row:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border

    # === Data Rows ===
    total_amount = 0
    for exp in expenses:
        ws.append([
            exp.date.strftime("%Y-%m-%d"),
            exp.category.name if exp.category else '',
            exp.description,
            float(exp.amount)
        ])
        total_amount += float(exp.amount)

    # === Total Row ===
    total_row = ws.max_row + 1
    ws[f"C{total_row}"] = "Total:"
    ws[f"C{total_row}"].font = Font(bold=True)
    ws[f"D{total_row}"] = total_amount
    ws[f"D{total_row}"].number_format = '"$"#,##0.00'

    # === Styling all data cells ===
    for row in ws.iter_rows(min_row=7, max_row=ws.max_row, min_col=1, max_col=4):
        for cell in row:
            cell.border = border
            if cell.column == 1:
                cell.alignment = center_alignment
            elif cell.column == 4:
                cell.number_format = '"$"#,##0.00'
                cell.alignment = Alignment(horizontal="right")

    # === Auto-adjust column widths ===
    for col in ws.iter_cols(min_row=6, max_row=ws.max_row, min_col=1, max_col=4):
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 4


    # === Response ===
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"Expenses_Report_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response
